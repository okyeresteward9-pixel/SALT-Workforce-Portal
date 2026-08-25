from gevent import monkey
monkey.patch_all()

from flask import Flask, render_template, request, redirect, session, jsonify, flash
import psycopg2
from psycopg2.extras import RealDictCursor
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
from flask_socketio import SocketIO, join_room  # pyright: ignore[reportMissingModuleSource]
from werkzeug.utils import secure_filename
import os
import time
from flask import send_from_directory
from datetime import date
from openpyxl import Workbook
from flask import send_file
import io
from io import BytesIO
from database import get_db, allowed_file, UPLOAD_FOLDER, ALLOWED_EXTENSIONS
from routes.chat import (
    chat_bp,
    register_chat_socketio
)
from request_workflow import register_request_workflow
from cloudinary import uploader
from cloudinary import api as cloudinary_api
import cloudinary
import cloudinary_config
import json
from push_notifications import get_vapid_public_key

app = Flask(__name__)


app.register_blueprint(chat_bp)

app.secret_key = os.environ.get(
    "SECRET_KEY",
    "fallback-secret"
)

# Session settings
app.permanent_session_lifetime = timedelta(days=30)

app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=True
)

socketio = SocketIO(
    app,
    cors_allowed_origins="*",
    async_mode="gevent"
)

register_chat_socketio(socketio)



def format_datetime(value):

    if not value:
        return {
            "date": "",
            "time": ""
        }

    dt = value if isinstance(value, datetime) else datetime.fromisoformat(value)

    return {
        "date": dt.strftime("%Y-%m-%d"),
        "time": dt.strftime("%H:%M")
    }

def build_comment_tree(comments):
    comment_map = {}
    tree = []

    for c in comments:
        c["children"] = []
        comment_map[c["id"]] = c

    for c in comments:
        parent_id = c["parent_comment_id"]

        if parent_id:
            parent = comment_map.get(parent_id)
            if parent:
                parent["children"].append(c)
        else:
            tree.append(c)

    return tree


# -------------------------
# Initialize database
# -------------------------
def init_db():
    conn = get_db()
    c = conn.cursor()

    # -------------------------
    # CREATE TABLES
    # -------------------------
    c.execute('''CREATE TABLE IF NOT EXISTS employees (
        id SERIAL PRIMARY KEY,
        name TEXT,
        email TEXT,
        password TEXT,
        role TEXT,
        phone TEXT,
        department TEXT,
        profile_pic TEXT,
        theme TEXT DEFAULT 'light'
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS attendance (
        id SERIAL PRIMARY KEY,
        employee_id INTEGER,
        clock_in TEXT,
        clock_out TEXT,
        latitude TEXT,
        longitude TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS tasks (
        id SERIAL PRIMARY KEY,
        title TEXT,
        description TEXT,
        assigned_to INTEGER,
        deadline TEXT,
        status TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS notifications (
        id SERIAL PRIMARY KEY,
        user_id INTEGER,
        message TEXT,
        is_read INTEGER DEFAULT 0,
        created_at TEXT
    )''')

    # -------------------------
    # WEB PUSH SUBSCRIPTIONS
    # -------------------------
    c.execute("""
    CREATE TABLE IF NOT EXISTS push_subscriptions (
        id SERIAL PRIMARY KEY,
        user_id INTEGER NOT NULL
            REFERENCES employees(id)
            ON DELETE CASCADE,
        endpoint TEXT NOT NULL UNIQUE,
        subscription JSONB NOT NULL,
        created_at TIMESTAMPTZ DEFAULT NOW()
    )
    """)

    # -------------------------
    # ANNOUNCEMENTS TABLE
    # -------------------------
    c.execute('''
    CREATE TABLE IF NOT EXISTS announcements (
        id SERIAL PRIMARY KEY,
        title TEXT NOT NULL,
        message TEXT NOT NULL,
        created_by INTEGER,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        audience TEXT DEFAULT 'everyone',
        file_name TEXT,
        file_path TEXT
    )
    ''')

    c.execute("""
    CREATE TABLE IF NOT EXISTS user_presence (

        user_id INTEGER PRIMARY KEY
            REFERENCES employees(id)
            ON DELETE CASCADE,

        online BOOLEAN DEFAULT FALSE,

        last_seen TIMESTAMP

    )
    """)

    # -------------------------
    # TASK COMMENTS TABLE
    # -------------------------
    c.execute('''
    CREATE TABLE IF NOT EXISTS task_comments (
        id SERIAL PRIMARY KEY,
        task_id INTEGER,
        sender_id INTEGER,
        sender_role TEXT,
        message TEXT,
        parent_comment_id INTEGER,
        created_at TEXT
    )''')

    # -------------------------
    # AUTO-FIX MISSING COLUMNS
    # -------------------------
    try:
        c.execute("ALTER TABLE employees ADD COLUMN phone TEXT")
    except:
        pass

    try:
        c.execute("ALTER TABLE employees ADD COLUMN department TEXT")
    except:
        pass

    try:
        c.execute("ALTER TABLE employees ADD COLUMN profile_pic TEXT")
    except:
        pass

    try:
        c.execute("ALTER TABLE employees ADD COLUMN theme TEXT DEFAULT 'light'")
    except:
        pass
    
    # -------------------------
    # ANNOUNCEMENT MIGRATIONS
    # -------------------------

    # Make sure PostgreSQL is not inside
    # an aborted transaction from an earlier
    # migration attempt.
    conn.rollback()

    # Add attachment path to existing databases
    c.execute("""
        ALTER TABLE announcements
        ADD COLUMN IF NOT EXISTS file_path TEXT
    """)

    # Add original attachment filename
    c.execute("""
        ALTER TABLE announcements
        ADD COLUMN IF NOT EXISTS file_name TEXT
    """)

    # Cloudinary asset identifiers for direct browser uploads.
    c.execute("""
        ALTER TABLE announcements
        ADD COLUMN IF NOT EXISTS cloudinary_public_id TEXT
    """)

    c.execute("""
        ALTER TABLE announcements
        ADD COLUMN IF NOT EXISTS cloudinary_resource_type TEXT
    """)

    c.execute("""
        ALTER TABLE announcements
        ADD COLUMN IF NOT EXISTS cloudinary_asset_id TEXT
    """)

    # Add announcement audience
    c.execute("""
        ALTER TABLE announcements
        ADD COLUMN IF NOT EXISTS audience TEXT
        DEFAULT 'everyone'
    """)

    # Add last-updated timestamp
    c.execute("""
        ALTER TABLE announcements
        ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP
    """)

    # -------------------------
    # FIX EXISTING RECORDS
    # -------------------------

    # Existing announcements should be
    # visible to everyone by default.
    c.execute("""
        UPDATE announcements
        SET audience = 'everyone'
        WHERE audience IS NULL
        OR audience = ''
    """)

    # Existing announcements were not edited,
    # so use created_at as their updated_at.
    c.execute("""
        UPDATE announcements
        SET updated_at = created_at::timestamp
        WHERE updated_at IS NULL
    """)

    try:
        c.execute("ALTER TABLE tasks ADD COLUMN created_by INTEGER")
    except:
        pass
    try:
        c.execute("UPDATE tasks SET status='Pending' WHERE status IS NULL")
    except:
        pass

    try:
        c. execute("ALTER TABLE tasks ADD COLUMN created_at TEXT;")
    except:
        pass

    try:
        c.execute("ALTER TABLE tasks ADD COLUMN completed_at TIMESTAMP;")
    except:
        pass

    try:
        c.execute("ALTER TABLE tasks ADD COLUMN carried_forward INTEGER DEFAULT 0;")
    except:
        pass
    try:
        c. execute("ALTER TABLE tasks ADD COLUMN original_deadline TEXT;")
    except:
        pass
    # Add note and reply columns safely
    try:
        c.execute("ALTER TABLE tasks ADD COLUMN note TEXT")
    except:
        pass

    try:
        c.execute("ALTER TABLE tasks ADD COLUMN reply TEXT")
    except:
        pass
    
    try:
        c.execute("ALTER TABLE tasks ADD COLUMN admin_reply TEXT")
    except:
        pass
    
    try:
        c.execute("ALTER TABLE tasks ADD COLUMN task_scope TEXT DEFAULT 'personal'")
    except:
        pass

    try:
        c.execute("ALTER TABLE employees ADD COLUMN position TEXT")
    except:
        pass

    try:
        c.execute("ALTER TABLE employees ADD COLUMN can_approve_requests BOOLEAN DEFAULT FALSE")
    except:
        pass

    try:
        c.execute("ALTER TABLE employees ADD COLUMN immediate_supervisor_id INTEGER")
    except:
        pass

    try:
        c.execute("ALTER TABLE employees ADD COLUMN signature_path TEXT")
    except:
        pass

    try:
        c.execute("ALTER TABLE employees ADD COLUMN profile_pic_public_id TEXT")
    except:
        pass
    # -------------------------
    # TASK VIEW VISIBILITY
    # -------------------------

    try:
        c.execute("""
            ALTER TABLE tasks
            ADD COLUMN admin_deleted BOOLEAN DEFAULT FALSE
        """)
    except:
        pass

    try:
        c.execute("""
            ALTER TABLE tasks
            ADD COLUMN employee_deleted BOOLEAN DEFAULT FALSE
        """)
    except:
        pass

    # Make sure existing tasks are visible
    try:
        c.execute("""
            UPDATE tasks
            SET admin_deleted = FALSE
            WHERE admin_deleted IS NULL
        """)
    except:
        pass

    try:
        c.execute("""
            UPDATE tasks
            SET employee_deleted = FALSE
            WHERE employee_deleted IS NULL
        """)
    except:
        pass

    try:
        c.execute("""
            ALTER TABLE task_comments
            ADD COLUMN visibility TEXT DEFAULT 'public'
        """)
    except:
        pass
    
    try:
        c.execute("""
            ALTER TABLE task_comments
            ADD COLUMN comment_type TEXT DEFAULT 'reply'
        """)
    except:
        pass


    conn.commit()
    conn.close()

# -------------------------
# Create admin if not exists
# -------------------------
def create_admin():
    conn = get_db()
    c = conn.cursor()

    admin_email = "admin@salt.com"
    hashed_password = generate_password_hash("Stgh2@&$%#3")

    c.execute(
        "SELECT id FROM employees WHERE email=%s",
        (admin_email,)
    )

    admin = c.fetchone()

    if not admin:
        c.execute("""
            INSERT INTO employees
            (name, email, password, role, theme)
            VALUES (%s,%s,%s,%s,%s)
        """, (
            "System Administrator",
            admin_email,
            hashed_password,
            "admin",
            "light"
        ))

        conn.commit()

    conn.close()

# -------------------------
# Routes
# -------------------------

@app.route('/')
def home():

    if 'user_id' in session:
        return redirect('/dashboard')

    return render_template("login.html")


@app.route('/login', methods=['POST'])
def login():

    email = request.form.get('email', '').strip()
    password = request.form.get('password', '')

    # Check Remember Me
    remember = request.form.get('remember') == '1'

    if not email or not password:
        return render_template(
            "login.html",
            error="Please enter your email and password."
        )

    conn = get_db()
    c = conn.cursor()

    c.execute(
        """
        SELECT *
        FROM employees
        WHERE email=%s
        """,
        (email,)
    )

    user = c.fetchone()
    conn.close()

    if user and check_password_hash(user['password'], password):

        # Clear old session data
        session.clear()

        # Store logged-in user
        session['user_id'] = user['id']
        session['name'] = user['name']
        session['role'] = user['role']

        # Remember Me
        if remember:
            session.permanent = True
        else:
            session.permanent = False

        return redirect('/dashboard')

    return render_template(
        "login.html",
        error="Invalid email or password."
    )

@app.route('/settings/profile', methods=['GET', 'POST'])
def profile_settings():

    if 'user_id' not in session:
        return redirect('/')

    conn = get_db()
    c = conn.cursor(cursor_factory=RealDictCursor)

    user_id = session['user_id']

    # ==========================================
    # UPDATE PROFILE
    # ==========================================

    if request.method == 'POST':

        # -------------------------
        # FORM DATA
        # -------------------------

        name = request.form.get('name', '').strip()
        phone = request.form.get('phone', '').strip()
        theme = request.form.get('theme', 'light')
        position = request.form.get('position', '').strip()

        new_password = request.form.get(
            'new_password',
            ''
        ).strip()

        confirm_password = request.form.get(
            'confirm_password',
            ''
        ).strip()

        current_password = request.form.get(
            'current_password',
            ''
        ).strip()


        # ==========================================
        # GET CURRENT USER
        # ==========================================

        c.execute(
            """
            SELECT
                password,
                department,
                profile_pic,
                profile_pic_public_id
            FROM employees
            WHERE id=%s
            """,
            (user_id,)
        )

        current_user = c.fetchone()

        if not current_user:
            conn.close()

            flash(
                "User account not found.",
                "error"
            )

            return redirect('/settings/profile')


        # ==========================================
        # PROFILE PICTURE
        # ==========================================

        profile_url = None
        profile_public_id = None

        file = request.files.get('image')

        if file and file.filename:

            filename = secure_filename(
                file.filename
            )

            if not filename:
                conn.close()

                flash(
                    "Invalid profile picture.",
                    "error"
                )

                return redirect('/settings/profile')


            ext = filename.rsplit(
                '.',
                1
            )[-1].lower()


            if ext not in ALLOWED_EXTENSIONS:

                conn.close()

                flash(
                    "Invalid file type. Only PNG, JPG and JPEG are allowed.",
                    "error"
                )

                return redirect('/settings/profile')


            try:

                # -------------------------
                # UPLOAD TO CLOUDINARY
                # -------------------------

                result = uploader.upload(
                    file,
                    folder="salt_portal/profiles",
                    public_id=f"user_{user_id}",
                    resource_type="image",
                    overwrite=True
                )


                profile_url = result.get(
                    "secure_url"
                )

                profile_public_id = result.get(
                    "public_id"
                )


                if not profile_url:

                    raise Exception(
                        "Cloudinary did not return a secure URL."
                    )


            except Exception as e:

                print(
                    "Cloudinary profile upload error:",
                    e
                )

                conn.close()

                flash(
                    "Unable to upload profile picture. Please try again.",
                    "error"
                )

                return redirect('/settings/profile')


        # ==========================================
        # DEPARTMENT CONTROL
        # ==========================================

        if session.get("role") == "admin":

            department = request.form.get(
                'department',
                ''
            ).strip()

        else:

            department = current_user.get(
                'department'
            )


        # ==========================================
        # PASSWORD VALIDATION
        # ==========================================

        update_password = False
        hashed_password = None


        if new_password:

            if not current_password:

                conn.close()

                flash(
                    "Please enter your current password.",
                    "error"
                )

                return redirect('/settings/profile')


            if new_password != confirm_password:

                conn.close()

                flash(
                    "Passwords do not match.",
                    "error"
                )

                return redirect('/settings/profile')


            stored_password = current_user.get(
                'password'
            )


            if not check_password_hash(
                stored_password,
                current_password
            ):

                conn.close()

                flash(
                    "Current password is incorrect.",
                    "error"
                )

                return redirect('/settings/profile')


            hashed_password = generate_password_hash(
                new_password
            )

            update_password = True


        # ==========================================
        # BUILD UPDATE QUERY
        # ==========================================

        query = """
            UPDATE employees
            SET
                name=%s,
                phone=%s,
                position=%s,
                department=%s,
                theme=%s
        """

        params = [
            name,
            phone,
            position,
            department,
            theme
        ]


        # -------------------------
        # PASSWORD
        # -------------------------

        if update_password:

            query += """
                ,
                password=%s
            """

            params.append(
                hashed_password
            )


        # -------------------------
        # PROFILE PICTURE
        # -------------------------

        if profile_url:

            query += """
                ,
                profile_pic=%s,
                profile_pic_public_id=%s
            """

            params.extend([
                profile_url,
                profile_public_id
            ])


        query += """
            WHERE id=%s
        """

        params.append(
            user_id
        )


        # ==========================================
        # SAVE DATABASE
        # ==========================================

        c.execute(
            query,
            tuple(params)
        )

        conn.commit()


        # ==========================================
        # DELETE OLD CLOUDINARY IMAGE
        # ==========================================
        #
        # IMPORTANT:
        # We use the same public_id with overwrite=True.
        # Therefore, Cloudinary has already replaced the
        # old image. DO NOT destroy it here.
        #

        if profile_url:

            print(
                "Profile image uploaded successfully:",
                profile_public_id
            )


        conn.close()


        # ==========================================
        # UPDATE SESSION
        # ==========================================

        session['name'] = name
        session['position'] = position


        flash(
            "Profile updated successfully!",
            "success"
        )

        return redirect(
            '/settings/profile'
        )


    # ==========================================
    # LOAD PROFILE
    # ==========================================

    c.execute(
        """
        SELECT *
        FROM employees
        WHERE id=%s
        """,
        (user_id,)
    )

    user = c.fetchone()

    conn.close()


    return render_template(
        "settings.html",
        user=user,
        role=session.get("role")
    )

@app.context_processor
def inject_user():

    if 'user_id' in session:

        conn = get_db()

        c = conn.cursor(cursor_factory=RealDictCursor)

        c.execute(
            """
            SELECT * 
            FROM employees 
            WHERE id=%s
            """,
            (session['user_id'],)
        )

        user = c.fetchone()

        conn.close()

        return dict(user=user)

    return dict(user=None)

# -------------------------
# WEB PUSH
# -------------------------

@app.route("/api/push/public-key")
def push_public_key():

    if 'user_id' not in session:
        return jsonify({"publicKey": ""}), 401

    return jsonify({
        "publicKey": get_vapid_public_key()
    })


@app.route("/api/push/subscribe", methods=["POST"])
def push_subscribe():

    if 'user_id' not in session:
        return jsonify({
            "success": False,
            "message": "Unauthorized"
        }), 401

    subscription = request.get_json(silent=True)

    if not subscription or not subscription.get("endpoint"):
        return jsonify({
            "success": False,
            "message": "Invalid push subscription."
        }), 400

    conn = get_db()
    c = conn.cursor()

    try:
        c.execute("""
            INSERT INTO push_subscriptions
            (
                user_id,
                endpoint,
                subscription
            )
            VALUES
            (
                %s,
                %s,
                %s::jsonb
            )
            ON CONFLICT (endpoint)
            DO UPDATE SET
                user_id = EXCLUDED.user_id,
                subscription = EXCLUDED.subscription
        """, (
            session['user_id'],
            subscription['endpoint'],
            json.dumps(subscription)
        ))

        conn.commit()

        return jsonify({
            "success": True
        })

    except Exception as e:
        conn.rollback()
        print("PUSH SUBSCRIPTION ERROR:", repr(e))

        return jsonify({
            "success": False,
            "message": "Unable to save push subscription."
        }), 500

    finally:
        conn.close()


# -------------------------
# Dashboard
# -------------------------
@app.route('/dashboard')
def dashboard():

    if 'user_id' not in session:
        return redirect('/')


    role = session.get('role')


    conn = get_db()
    c = conn.cursor()



    # -------------------------
    # Attendance logic
    # -------------------------

    today = datetime.now().date()
    today_str = today.strftime("%Y-%m-%d")


    c.execute("""
        SELECT clock_in, clock_out
        FROM attendance
        WHERE employee_id=%s
        AND DATE(clock_in)=DATE(%s)
        ORDER BY id DESC
        LIMIT 1
    """,
    (
        session['user_id'],
        today_str
    ))


    attendance_record = c.fetchone()


    clocked_in = False
    clocked_out = False


    if attendance_record:

        clocked_in = True

        if attendance_record['clock_out'] is not None:
            clocked_out = True



    # -------------------------
    # Dashboard stats
    # -------------------------

    total_employees = 0
    present_today = 0
    absent_today = 0
    clockins_today = 0


    if role == 'admin':

        c.execute(
            "SELECT COUNT(*) AS count FROM employees"
        )

        total_employees = c.fetchone()['count']


        c.execute("""
            SELECT COUNT(DISTINCT employee_id) AS count
            FROM attendance
            WHERE DATE(clock_in)=%s
        """,
        (today_str,))


        present_today = c.fetchone()['count']


        absent_today = total_employees - present_today



        c.execute("""
            SELECT COUNT(*) AS count
            FROM attendance
            WHERE DATE(clock_in)=%s
        """,
        (today_str,))


        clockins_today = c.fetchone()['count']


    else:


        c.execute("""
            SELECT COUNT(*) AS count
            FROM attendance
            WHERE employee_id=%s
            AND DATE(clock_in)=%s
        """,
        (
            session['user_id'],
            today_str
        ))


        clockins_today = c.fetchone()['count']



    # -------------------------
    # Weekly chart data
    # -------------------------

    week_data = []
    labels = []


    for i in range(7):

        day = today - timedelta(days=6-i)

        c.execute("""
            SELECT COUNT(DISTINCT employee_id) AS count
            FROM attendance
            WHERE DATE(clock_in)=%s
        """,
        (
            day.strftime("%Y-%m-%d"),
        ))


        week_data.append(
            c.fetchone()['count']
        )

        labels.append(
            day.strftime("%a")
        )



    # -------------------------
    # Working employees
    # -------------------------

    if role == 'admin':

        c.execute("""
            SELECT 
                employees.name,
                attendance.clock_in,
                attendance.latitude,
                attendance.longitude
            FROM attendance

            JOIN employees
            ON attendance.employee_id = employees.id

            WHERE DATE(clock_in)=%s
            AND clock_out IS NULL
        """,
        (today_str,))


        working_employees = c.fetchall()


    else:

        working_employees = []



    # -------------------------
    # Format clock-in time
    # -------------------------

    formatted_employees = []


    for emp in working_employees:


        emp = dict(emp)


        try:

            dt = emp['clock_in']

            if isinstance(dt, str):
                dt = datetime.fromisoformat(dt)


            emp['clock_in'] = dt.strftime("%H:%M")


        except Exception:
            pass


        formatted_employees.append(emp)



    working_employees = formatted_employees



    # -------------------------
    # Task stats
    # -------------------------

    if role == 'admin':

        c.execute("""
            SELECT COUNT(*) AS count
            FROM tasks
            WHERE status='Completed'
        """)

        tasks_completed = c.fetchone()['count']


        c.execute("""
            SELECT COUNT(*) AS count
            FROM tasks
            WHERE status!='Completed'
        """)

        tasks_pending = c.fetchone()['count']


    else:


        c.execute("""
            SELECT COUNT(*) AS count
            FROM tasks
            WHERE assigned_to=%s
            AND status='Completed'
        """,
        (session['user_id'],))


        tasks_completed = c.fetchone()['count']



        c.execute("""
            SELECT COUNT(*) AS count
            FROM tasks
            WHERE assigned_to=%s
            AND status!='Completed'
        """,
        (session['user_id'],))


        tasks_pending = c.fetchone()['count']



    # -------------------------
    # Notifications
    # -------------------------

    notification_count = get_notification_count(
        session['user_id']
    )

    # -------------------------
    # Request workflow counters
    # -------------------------
    request_pending_approvals = 0
    request_my_pending = 0

    try:
        c.execute("""
            SELECT COUNT(*) AS count
            FROM request_steps
            WHERE approver_id=%s AND status='pending'
        """, (session['user_id'],))
        request_pending_approvals = c.fetchone()['count']

        c.execute("""
            SELECT COUNT(*) AS count
            FROM requests
            WHERE requester_id=%s
              AND status NOT IN ('completed','rejected')
        """, (session['user_id'],))
        request_my_pending = c.fetchone()['count']
    except Exception:
        request_pending_approvals = 0
        request_my_pending = 0



    # -------------------------
    # Announcements
    # -------------------------

    if role == 'admin':

        # Admins can see:
        # - Everyone announcements
        # - Staff announcements
        # - Admin announcements

        c.execute("""
            SELECT
                announcements.id,
                announcements.title,
                announcements.message,
                announcements.created_at,
                announcements.updated_at,
                announcements.audience,
                announcements.file_name,
                announcements.file_path,

                COALESCE(
                    employees.name,
                    'Unknown'
                ) AS created_by_name

            FROM announcements

            LEFT JOIN employees
                ON announcements.created_by = employees.id

            WHERE announcements.audience IN (
                'everyone',
                'staff',
                'admin'
            )

            ORDER BY announcements.created_at DESC

            LIMIT 5
        """)

    else:

        # Staff can see:
        # - Everyone announcements
        # - Staff announcements
        #
        # Admin-only announcements are hidden.

        c.execute("""
            SELECT
                announcements.id,
                announcements.title,
                announcements.message,
                announcements.created_at,
                announcements.updated_at,
                announcements.audience,
                announcements.file_name,
                announcements.file_path,

                COALESCE(
                    employees.name,
                    'Unknown'
                ) AS created_by_name

            FROM announcements

            LEFT JOIN employees
                ON announcements.created_by = employees.id

            WHERE announcements.audience IN (
                'everyone',
                'staff'
            )

            ORDER BY announcements.created_at DESC

            LIMIT 5
        """)

    latest_announcements = c.fetchall()


    c.close()
    conn.close()



    # -------------------------
    # Employee Locations
    # -------------------------

    employee_locations = []


    for emp in working_employees:

        try:

            lat_f = float(emp['latitude'])
            lon_f = float(emp['longitude'])


        except (TypeError, ValueError):

            continue


        employee_locations.append({

            "name": emp['name'],
            "time": emp['clock_in'],
            "lat": lat_f,
            "lon": lon_f

        })



    return render_template(
        "dashboard.html",

        name=session['name'],
        role=role,

        total_employees=total_employees,
        present_today=present_today,
        absent_today=absent_today,
        clockins_today=clockins_today,

        notification_count=notification_count,

        request_pending_approvals=request_pending_approvals,
        request_my_pending=request_my_pending,

        clocked_in=clocked_in,
        clocked_out=clocked_out,

        week_data=week_data,
        labels=labels,

        working_employees=working_employees,

        tasks_completed=tasks_completed,
        tasks_pending=tasks_pending,

        latest_announcements=latest_announcements,

        employee_locations=employee_locations
    )
# -------------------------
# Clock-in
# -------------------------
@app.route('/clockin', methods=['POST'])
def clockin():

    if 'user_id' not in session:
        return redirect('/')


    lat = request.form['latitude']
    lon = request.form['longitude']


    conn = get_db()
    c = conn.cursor()


    now = datetime.now()


    c.execute("""
        INSERT INTO attendance 
        (
            employee_id,
            clock_in,
            latitude,
            longitude
        )
        VALUES (%s, %s, %s, %s)
    """,
    (
        session['user_id'],
        now,
        lat,
        lon
    ))


    conn.commit()
    conn.close()


    return redirect('/dashboard')
# -------------------------
# Clock-out
# -------------------------
@app.route('/clockout', methods=['POST'])
def clockout():

    if 'user_id' not in session:
        return redirect('/')


    conn = get_db()
    c = conn.cursor(cursor_factory=RealDictCursor)


    c.execute("""
        UPDATE attendance
        SET clock_out=%s
        WHERE employee_id=%s
        AND DATE(clock_in)=CURRENT_DATE
        AND clock_out IS NULL
    """,
    (
        datetime.now(),
        session['user_id']
    ))


    conn.commit()
    conn.close()


    return redirect('/dashboard')

@app.route("/service-worker.js")
def service_worker():
    return send_from_directory(
        os.path.join(app.root_path, "static", "js"),
        "service-worker.js",
        mimetype="application/javascript"
    )

# -------------------------
# Notifications helper
# -------------------------
def create_notification(user_id, message, created_at=None):
    """
    Create a persistent notification and immediately push it
    to the user's Socket.IO room.

    This is the single notification entry point for the portal.
    It is intentionally generic so tasks, announcements,
    comments, workflow events, and future features can all use
    the same notification system.
    """
    if not user_id or not message:
        return None

    created_at = created_at or datetime.now()

    conn = get_db()
    c = conn.cursor(cursor_factory=RealDictCursor)

    try:
        c.execute("""
            INSERT INTO notifications
            (
                user_id,
                message,
                is_read,
                created_at
            )
            VALUES
            (
                %s,
                %s,
                FALSE,
                %s
            )
            RETURNING id
        """, (
            int(user_id),
            message,
            created_at
        ))

        row = c.fetchone()
        notification_id = row["id"]

        conn.commit()

        socketio.emit(
            "new_notification",
            {
                "id": notification_id,
                "message": message,
                "created_at": created_at.strftime("%Y-%m-%d %H:%M:%S"),
                "is_read": False
            },
            room=f"user_{int(user_id)}"
        )

        # Browser/desktop push is deliberately best-effort.
        # If push is unavailable, the normal SALT notification
        # system above continues to work normally.
        try:
            from push_notifications import send_web_push

            c.execute("""
                SELECT subscription
                FROM push_subscriptions
                WHERE user_id=%s
            """, (int(user_id),))

            push_rows = c.fetchall()

            for push_row in push_rows:
                subscription = push_row["subscription"]

                if isinstance(subscription, str):
                    subscription = json.loads(subscription)

                send_web_push(
                    subscription=subscription,
                    title="SALT Portal",
                    body=message,
                    url="/dashboard",
                    tag=f"salt-notification-{notification_id}"
                )

        except Exception as push_error:
            print(
                "DESKTOP PUSH SKIPPED:",
                repr(push_error)
            )

        return notification_id

    except Exception as e:
        conn.rollback()
        print("CREATE NOTIFICATION ERROR:", repr(e))
        return None

    finally:
        conn.close()


def notify_users(user_ids, message):
    """
    Send the same notification to multiple users.
    Duplicate user IDs are ignored.
    """
    sent = []

    for user_id in set(user_ids or []):
        notification_id = create_notification(
            user_id,
            message
        )

        if notification_id is not None:
            sent.append(notification_id)

    return sent


def notify_announcement(audience, message, exclude_user_id=None):
    """
    Notify users according to an announcement audience:
      - everyone -> all employees
      - staff    -> staff users only
      - admin    -> admin users only
    """
    conn = get_db()
    c = conn.cursor()

    try:
        if audience == "admin":
            c.execute("""
                SELECT id
                FROM employees
                WHERE role='admin'
            """)
        elif audience == "staff":
            c.execute("""
                SELECT id
                FROM employees
                WHERE role <> 'admin'
            """)
        else:
            c.execute("""
                SELECT id
                FROM employees
            """)

        user_ids = [
            row["id"]
            for row in c.fetchall()
            if row["id"] != exclude_user_id
        ]

    finally:
        conn.close()

    return notify_users(
        user_ids,
        message
    )


def get_notification_count(user_id):

    conn = get_db()
    c = conn.cursor()

    c.execute("""
    SELECT COUNT(*) AS total
    FROM notifications
    WHERE user_id=%s AND is_read=FALSE
    """, (user_id,))

    count = c.fetchone()["total"]

    conn = get_db()
    c = conn.cursor()

    c.execute("""
    SELECT COUNT(*) AS total
    FROM notifications
    WHERE user_id=%s AND is_read=FALSE
    """, (user_id,))

    count = c.fetchone()["total"]


    conn.close()

    return count

# -------------------------
# Get notifications (AJAX)
# -------------------------
@app.route('/api/notifications')
def api_notifications():

    if 'user_id' not in session:
        return jsonify([])


    conn = get_db()

    c = conn.cursor()


    c.execute("""
        SELECT id, message, created_at, is_read
        FROM notifications
        WHERE user_id=%s
        ORDER BY created_at DESC
        LIMIT 10
    """,
    (session['user_id'],))


    notes = c.fetchall()


    conn.close()


    return jsonify(notes)


# -------------------------
# Mark all notifications read
# -------------------------
@app.route('/api/notifications/mark-read', methods=['POST'])
def mark_notifications_read():

    if 'user_id' not in session:
        return jsonify({"success": False})


    conn = get_db()
    c = conn.cursor()


    c.execute("""
        UPDATE notifications
        SET is_read=TRUE
        WHERE user_id=%s
    """,
    (session['user_id'],))


    conn.commit()
    conn.close()


    return jsonify({"success": True})

# -------------------------
# Clear all notifications
# -------------------------
@app.route('/api/notifications/clear-all', methods=['POST'])
def clear_all_notifications():

    if 'user_id' not in session:
        return jsonify({
            "success": False,
            "message": "Unauthorized"
        }), 401

    conn = get_db()
    c = conn.cursor()

    try:

        c.execute("""
            DELETE FROM notifications
            WHERE user_id=%s
        """, (
            session['user_id'],
        ))

        deleted_count = c.rowcount

        conn.commit()

        return jsonify({
            "success": True,
            "deleted": deleted_count
        })

    except Exception as e:

        conn.rollback()

        print(
            "CLEAR NOTIFICATIONS ERROR:",
            e
        )

        return jsonify({
            "success": False,
            "message": "Unable to clear notifications."
        }), 500

    finally:

        conn.close()

# -------------------------
# Clear all notifications
def get_attendance_date_range(period, start_date_value, end_date_value):
    """
    Return (start_datetime, end_datetime) for the selected filter.

    end_datetime is exclusive.

    Custom dates are treated as whole calendar days:
        From 2026-08-01
        To   2026-08-10

    means:
        >= 2026-08-01 00:00:00
        <  2026-08-11 00:00:00
    """

    today = date.today()

    # ------------------------------------------------------------
    # QUICK FILTERS
    # ------------------------------------------------------------

    if period == "today":

        start_date = today
        end_date = today

    elif period == "yesterday":

        start_date = today - timedelta(days=1)
        end_date = start_date

    elif period == "this_week":

        # Monday -> Sunday
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)

    elif period == "last_week":

        this_week_start = today - timedelta(days=today.weekday())

        start_date = this_week_start - timedelta(days=7)
        end_date = this_week_start - timedelta(days=1)

    elif period == "this_month":

        start_date = today.replace(day=1)

        if today.month == 12:
            next_month = date(
                today.year + 1,
                1,
                1
            )
        else:
            next_month = date(
                today.year,
                today.month + 1,
                1
            )

        end_date = next_month - timedelta(days=1)

    elif period == "last_month":

        first_this_month = today.replace(day=1)

        last_day_last_month = (
            first_this_month - timedelta(days=1)
        )

        start_date = last_day_last_month.replace(day=1)
        end_date = last_day_last_month

    # ------------------------------------------------------------
    # CUSTOM DATE RANGE
    # ------------------------------------------------------------

    elif start_date_value or end_date_value:

        try:

            start_date = (
                datetime.strptime(
                    start_date_value,
                    "%Y-%m-%d"
                ).date()
                if start_date_value
                else None
            )

            end_date = (
                datetime.strptime(
                    end_date_value,
                    "%Y-%m-%d"
                ).date()
                if end_date_value
                else None
            )

        except (
            ValueError,
            TypeError
        ):

            start_date = None
            end_date = None

    else:

        start_date = None
        end_date = None

    # ------------------------------------------------------------
    # CONVERT DATES TO DATETIME BOUNDARIES
    # ------------------------------------------------------------

    start_datetime = None
    end_datetime = None

    if start_date:
        start_datetime = datetime.combine(
            start_date,
            datetime.min.time()
        )

    if end_date:
        # End is exclusive, therefore use the next day.
        end_datetime = datetime.combine(
            end_date + timedelta(days=1),
            datetime.min.time()
        )

    return start_datetime, end_datetime


@app.route('/attendance')
def attendance():

    if 'user_id' not in session:
        return redirect('/')


    conn = get_db()

    c = conn.cursor()


    role = session.get('role')


    # ============================================================
    # FILTER VALUES
    # ============================================================

    period = request.args.get(
        "period",
        ""
    ).strip()

    employee_id = request.args.get(
        "employee_id",
        ""
    ).strip()

    status = request.args.get(
        "status",
        ""
    ).strip().lower()

    start_date_value = request.args.get(
        "start_date",
        ""
    ).strip()

    end_date_value = request.args.get(
        "end_date",
        ""
    ).strip()


    # ============================================================
    # DATE RANGE
    # ============================================================

    start_datetime, end_datetime = (
        get_attendance_date_range(
            period,
            start_date_value,
            end_date_value
        )
    )


    # ============================================================
    # BASE QUERY
    # ============================================================

    query = """

        SELECT

            employees.name,

            attendance.clock_in,

            attendance.clock_out

        FROM attendance

        JOIN employees

            ON attendance.employee_id =
               employees.id

        WHERE 1=1

    """


    params = []


    # ============================================================
    # STAFF SECURITY
    #
    # Staff can ONLY see their own attendance.
    # ============================================================

    if role != 'admin':

        query += """

            AND attendance.employee_id = %s

        """

        params.append(
            session['user_id']
        )


    # ============================================================
    # ADMIN → EMPLOYEE FILTER
    # ============================================================

    elif employee_id:

        try:

            selected_employee_id = int(
                employee_id
            )

            query += """

                AND attendance.employee_id = %s

            """

            params.append(
                selected_employee_id
            )

        except (
            ValueError,
            TypeError
        ):

            employee_id = ""


    # ============================================================
    # DATE FILTER
    # ============================================================

    if start_datetime:

        query += """

            AND attendance.clock_in::timestamp >= %s

        """

        params.append(
            start_datetime
        )


    if end_datetime:

        query += """

            AND attendance.clock_in::timestamp < %s

        """

        params.append(
            end_datetime
        )


    # ============================================================
    # STATUS FILTER
    # ============================================================

    if status == "active":

        query += """

            AND attendance.clock_out IS NULL

        """

    elif status == "completed":

        query += """

            AND attendance.clock_out IS NOT NULL

        """


    # ============================================================
    # ORDER
    # ============================================================

    query += """

        ORDER BY attendance.clock_in DESC

    """


    c.execute(
        query,
        params
    )


    raw_records = c.fetchall()


    # ============================================================
    # ADMIN EMPLOYEE LIST
    # ============================================================

    employees = []


    if role == 'admin':

        c.execute("""

            SELECT
                id,
                name

            FROM employees

            ORDER BY name ASC

        """)

        employees = c.fetchall()


    conn.close()


    # ============================================================
    # FORMAT DATA
    # ============================================================

    records = []


    for r in raw_records:

        clock_in = format_datetime(
            r["clock_in"]
        )

        clock_out = format_datetime(
            r["clock_out"]
        )


        records.append({

            "name":
                r["name"]
                if role == 'admin'
                else None,

            "clock_in_date":
                clock_in["date"],

            "clock_in_time":
                clock_in["time"],

            "clock_out_date":
                clock_out["date"],

            "clock_out_time":
                clock_out["time"],

            "is_active":
                r["clock_out"] is None

        })


    # ============================================================
    # RENDER
    # ============================================================

    return render_template(

        "attendance.html",

        records=records,

        role=role,

        employees=employees,

        selected_employee=employee_id,

        status=status,

        period=period,

        start_date=start_date_value,

        end_date=end_date_value

    )


# ============================================================
# EXPORT ATTENDANCE
# ============================================================

@app.route('/export/attendance')
def export_attendance():

    if 'user_id' not in session:
        return redirect('/')


    role = session.get('role')


    conn = get_db()

    c = conn.cursor()


    # ============================================================
    # SAME FILTERS AS ATTENDANCE PAGE
    # ============================================================

    period = request.args.get(
        "period",
        ""
    ).strip()

    employee_id = request.args.get(
        "employee_id",
        ""
    ).strip()

    status = request.args.get(
        "status",
        ""
    ).strip().lower()

    start_date_value = request.args.get(
        "start_date",
        ""
    ).strip()

    end_date_value = request.args.get(
        "end_date",
        ""
    ).strip()


    start_datetime, end_datetime = (
        get_attendance_date_range(
            period,
            start_date_value,
            end_date_value
        )
    )


    # ============================================================
    # BASE QUERY
    # ============================================================

    query = """

        SELECT

            employees.name,

            attendance.clock_in,

            attendance.clock_out

        FROM attendance

        JOIN employees

            ON attendance.employee_id =
               employees.id

        WHERE 1=1

    """


    params = []


    # ============================================================
    # STAFF SECURITY
    # ============================================================

    if role != 'admin':

        query += """

            AND attendance.employee_id = %s

        """

        params.append(
            session['user_id']
        )


    # ============================================================
    # ADMIN → EMPLOYEE FILTER
    # ============================================================

    elif employee_id:

        try:

            selected_employee_id = int(
                employee_id
            )

            query += """

                AND attendance.employee_id = %s

            """

            params.append(
                selected_employee_id
            )

        except (
            ValueError,
            TypeError
        ):

            employee_id = ""


    # ============================================================
    # DATE FILTER
    # ============================================================

    if start_datetime:

        query += """

            AND attendance.clock_in::timestamp >= %s

        """

        params.append(
            start_datetime
        )


    if end_datetime:

        query += """

            AND attendance.clock_in::timestamp < %s

        """

        params.append(
            end_datetime
        )


    # ============================================================
    # STATUS FILTER
    # ============================================================

    if status == "active":

        query += """

            AND attendance.clock_out IS NULL

        """

    elif status == "completed":

        query += """

            AND attendance.clock_out IS NOT NULL

        """


    query += """

        ORDER BY attendance.clock_in DESC

    """


    c.execute(
        query,
        params
    )


    records = c.fetchall()


    conn.close()


    # ============================================================
    # CREATE EXCEL
    # ============================================================

    wb = Workbook()

    ws = wb.active

    ws.title = "Attendance"


    # ============================================================
    # HEADERS
    # ============================================================

    if role == 'admin':

        ws.append([

            "Employee",
            "Clock In",
            "Clock Out",
            "Status"

        ])

    else:

        ws.append([

            "Clock In",
            "Clock Out",
            "Status"

        ])


    # ============================================================
    # DATA
    # ============================================================

    for r in records:

        clock_in = r["clock_in"]

        clock_out = r["clock_out"]


        status_text = (

            "Active"

            if clock_out is None

            else "Completed"

        )


        if role == 'admin':

            ws.append([

                r["name"],

                clock_in,

                clock_out
                or "Still working",

                status_text

            ])

        else:

            ws.append([

                clock_in,

                clock_out
                or "Still working",

                status_text

            ])


    # ============================================================
    # BASIC EXCEL FORMATTING
    # ============================================================

    for cell in ws[1]:

        cell.font = cell.font.copy(
            bold=True
        )


    ws.freeze_panes = "A2"


    for column in ws.columns:

        max_length = 0

        column_letter = column[0].column_letter

        for cell in column:

            value = cell.value

            if value is not None:

                max_length = max(
                    max_length,
                    len(str(value))
                )

        ws.column_dimensions[
            column_letter
        ].width = min(
            max_length + 3,
            35
        )


    # ============================================================
    # SAVE TO MEMORY
    # ============================================================

    file_stream = io.BytesIO()

    wb.save(
        file_stream
    )

    file_stream.seek(0)


    return send_file(

        file_stream,

        as_attachment=True,

        download_name="attendance.xlsx",

        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )

    )

@app.route('/admin/tasks')
def admin_tasks():

    if 'role' not in session or session['role'] != 'admin':
        return "Access Denied"


    conn = get_db()

    c = conn.cursor(
        cursor_factory=RealDictCursor
    )


    status = request.args.get("status")

    # Employee filter
    employee_id = request.args.get("employee_id")


    # =========================================================
    # LOAD EMPLOYEES FOR FILTER DROPDOWN
    # =========================================================

    c.execute("""
        SELECT
            id,
            name
        FROM employees
        ORDER BY name ASC
    """)

    employees = c.fetchall()


    # =========================================================
    # LOAD ALL TASKS VISIBLE TO ADMIN
    #
    # We do NOT filter by task_scope.
    #
    # Self tasks
    # Assigned tasks
    # Admin-created tasks
    # Employee-created tasks
    #
    # ALL can appear on the Admin Task Board.
    #
    # admin_deleted only controls whether the admin has
    # removed the task from THEIR view.
    # =========================================================

    base_query = """

        SELECT

            tasks.id,
            tasks.title,
            tasks.description,
            tasks.note,
            tasks.reply,
            tasks.admin_reply,
            tasks.deadline,
            tasks.status,
            tasks.created_by,
            tasks.assigned_to,
            tasks.carried_forward,
            tasks.task_scope,
            tasks.created_at,

            tasks.admin_deleted,
            tasks.employee_deleted,

            employees.name AS employee_name

        FROM tasks

        LEFT JOIN employees

            ON tasks.assigned_to = employees.id

        WHERE COALESCE(
            tasks.admin_deleted,
            FALSE
        ) = FALSE

    """


    params = []


    # =========================================================
    # EMPLOYEE NAME FILTER
    # =========================================================

    if employee_id:

        base_query += """

            AND tasks.assigned_to = %s

        """

        params.append(employee_id)


    # =========================================================
    # STATUS FILTER
    # =========================================================

    if status:

        base_query += """

            AND tasks.status = %s

        """

        params.append(status)


    # =========================================================
    # ORDER
    # =========================================================

    base_query += """

        ORDER BY tasks.id DESC

    """


    c.execute(
        base_query,
        params
    )


    tasks = c.fetchall()


    # =========================================================
    # NORMALIZE DEADLINES
    # =========================================================

    for task in tasks:

        deadline = task.get("deadline")


        if deadline:

            if isinstance(
                deadline,
                datetime
            ):

                task["deadline_date"] = (
                    deadline.date()
                )


            elif isinstance(
                deadline,
                date
            ):

                task["deadline_date"] = deadline


            elif isinstance(
                deadline,
                str
            ):

                try:

                    task["deadline_date"] = (
                        datetime.strptime(
                            deadline[:10],
                            "%Y-%m-%d"
                        ).date()
                    )

                except (
                    ValueError,
                    TypeError
                ):

                    task["deadline_date"] = None


            else:

                task["deadline_date"] = None

        else:

            task["deadline_date"] = None


    # =========================================================
    # COMMENTS / DISCUSSIONS
    # =========================================================

    for task in tasks:

        c.execute("""

            SELECT *

            FROM task_comments

            WHERE task_id = %s

            ORDER BY created_at ASC

        """,
        (
            task["id"],
        ))


        comments = c.fetchall()


        task["comments"] = (
            build_comment_tree(
                comments
            )
        )


    # =========================================================
    # STATISTICS
    #
    # Statistics follow the selected employee filter.
    # Status buttons continue to filter the task list.
    #
    # We intentionally DO NOT apply the status filter here,
    # so the three cards show Pending/In Progress/Completed
    # for the selected employee.
    # =========================================================

    stats_params = []


    stats_where = """

        WHERE COALESCE(
            admin_deleted,
            FALSE
        ) = FALSE

    """


    if employee_id:

        stats_where += """

            AND assigned_to = %s

        """

        stats_params.append(employee_id)


    c.execute(f"""

        SELECT COUNT(*) AS count

        FROM tasks

        {stats_where}

        AND status = 'Pending'

    """,
    stats_params)


    pending_count = (
        c.fetchone()["count"]
    )


    c.execute(f"""

        SELECT COUNT(*) AS count

        FROM tasks

        {stats_where}

        AND status = 'In Progress'

    """,
    stats_params)


    progress_count = (
        c.fetchone()["count"]
    )


    c.execute(f"""

        SELECT COUNT(*) AS count

        FROM tasks

        {stats_where}

        AND status = 'Completed'

    """,
    stats_params)


    completed_count = (
        c.fetchone()["count"]
    )


    conn.close()


    # =========================================================
    # TODAY
    # =========================================================

    today = date.today()


    # =========================================================
    # RENDER
    # =========================================================

    return render_template(

        "admin_tasks.html",

        tasks=tasks,

        employees=employees,

        selected_employee=employee_id,

        selected_status=status,

        pending_count=pending_count,

        progress_count=progress_count,

        completed_count=completed_count,

        today=today

    )


@app.route('/admin/reply_task/<int:id>', methods=['POST'])
def admin_reply_task(id):

    if 'role' not in session or session['role'] != 'admin':
        return jsonify({
            "status": "error",
            "message": "Access Denied"
        }), 403

    message = request.form.get(
        'admin_reply',
        ''
    ).strip()

    parent_comment_id = request.form.get(
        'parent_comment_id'
    )

    if not message:
        return jsonify({
            "status": "error",
            "message": "Reply cannot be empty."
        }), 400

    conn = get_db()

    c = conn.cursor(
        cursor_factory=RealDictCursor
    )

    try:

        # =====================================================
        # FIND TASK
        # =====================================================

        c.execute("""
            SELECT
                id,
                assigned_to,
                title
            FROM tasks
            WHERE id = %s
        """, (id,))

        task_owner = c.fetchone()

        if not task_owner:

            return jsonify({
                "status": "error",
                "message": "Task not found."
            }), 404


        # =====================================================
        # INSERT ADMIN REPLY
        # =====================================================

        c.execute("""
            INSERT INTO task_comments
            (
                task_id,
                sender_id,
                sender_role,
                message,
                parent_comment_id,
                visibility,
                created_at
            )
            VALUES
            (
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s
            )
            RETURNING id
        """,
        (
            id,
            session['user_id'],
            'admin',
            message,
            parent_comment_id,
            'public',
            datetime.now()
        ))

        new_comment = c.fetchone()

        comment_id = (
            new_comment["id"]
            if new_comment
            else None
        )


        # =====================================================
        # SAVE
        # =====================================================

        conn.commit()


        # =====================================================
        # NOTIFY EMPLOYEE
        # =====================================================

        employee_id = task_owner["assigned_to"]

        if employee_id:

            try:

                create_notification(
                    employee_id,
                    f"Admin replied to your task: {task_owner['title']}"
                )

            except Exception as notification_error:

                print(
                    "ADMIN REPLY NOTIFICATION ERROR:",
                    notification_error
                )


        # =====================================================
        # JSON RESPONSE
        # =====================================================

        return jsonify({

            "status": "success",

            "message":
                "Reply posted successfully.",

            "task_id": id,

            "comment_id": comment_id

        }), 200


    except Exception as e:

        conn.rollback()

        print(
            "ADMIN REPLY TASK ERROR:",
            repr(e)
        )

        return jsonify({

            "status": "error",

            "message":
                "Unable to post the reply."

        }), 500


    finally:

        conn.close()

@app.route(
    '/delete_task/<int:task_id>',
    methods=['POST']
)
def delete_task_route(task_id):

    # =====================================================
    # AUTHENTICATION
    # =====================================================

    if 'user_id' not in session:

        return jsonify({

            "status": "error",

            "message": "Please log in."

        }), 401


    # =====================================================
    # ADMIN ONLY
    # =====================================================

    if session.get('role') != 'admin':

        return jsonify({

            "status": "error",

            "message": "Access Denied"

        }), 403


    conn = get_db()

    c = conn.cursor(
        cursor_factory=RealDictCursor
    )


    try:

        # =================================================
        # CHECK TASK
        # =================================================

        c.execute("""

            SELECT
                id,
                title,
                admin_deleted

            FROM tasks

            WHERE id = %s

        """,
        (
            task_id,
        ))


        task = c.fetchone()


        if not task:

            conn.close()

            return jsonify({

                "status": "error",

                "message": "Task not found."

            }), 404


        # =================================================
        # HIDE FROM ADMIN VIEW ONLY
        #
        # DO NOT DELETE THE TASK.
        # DO NOT DELETE COMMENTS.
        # =================================================

        c.execute("""

            UPDATE tasks

            SET admin_deleted = TRUE

            WHERE id = %s

        """,
        (
            task_id,
        ))


        conn.commit()

        conn.close()


        # =================================================
        # AJAX RESPONSE
        # =================================================

        return jsonify({

            "status": "success",

            "message":
                "Task removed from the Admin Task Board.",

            "task_id": task_id

        }), 200


    except Exception as e:

        conn.rollback()

        conn.close()


        print(
            "ADMIN DELETE TASK ERROR:",
            e
        )


        return jsonify({

            "status": "error",

            "message":
                "Unable to remove task."

        }), 500

@app.route('/admin/assign_task', methods=['GET','POST'])
def assign_task():

    if 'role' not in session or session['role'] != 'admin':
        return "Access Denied"


    conn = get_db()

    c = conn.cursor(cursor_factory=RealDictCursor)



    if request.method == 'POST':

        title = request.form['title']

        description = request.form['description']

        note = request.form['note']

        assigned_to = request.form['assigned_to']

        deadline = request.form.get('deadline') or None



        created_at = datetime.now()



        c.execute("""
            INSERT INTO tasks
            (
                title,
                description,
                note,
                assigned_to,
                deadline,
                status,
                created_by,
                task_scope,
                created_at
            )

            VALUES
            (
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s
            )

        """,
        (
            title,
            description,
            note,
            assigned_to,
            deadline,
            "Pending",
            session['user_id'],
            "admin_board",
            created_at
        ))



        conn.commit()

        conn.close()

        # 🔔 Persistent + real-time notification
        create_notification(
            assigned_to,
            f"New Task Assigned: {title}"
        )

        return redirect('/admin/tasks')



    # Load employees

    c.execute("""
        SELECT id, name
        FROM employees
        ORDER BY name ASC
    """)


    employees = c.fetchall()


    conn.close()


    return render_template(
        "assign_task.html",
        employees=employees
    )

@app.route('/admin/employees')
def admin_employees():

    if 'role' not in session or session['role'] != 'admin':
        return "Access Denied", 403

    conn = get_db()
    c = conn.cursor(cursor_factory=RealDictCursor)

    try:
        c.execute("""
            SELECT
                id,
                name,
                email,
                role,
                phone,
                department,
                position,
                can_approve_requests,
                immediate_supervisor_id,
                signature_path,
                profile_pic,
                theme,
                profile_pic_public_id
            FROM employees
            ORDER BY id DESC
        """)

        employees = c.fetchall()

    finally:
        conn.close()

    return render_template(
        "employees.html",
        employees=employees
    )

@app.route("/admin/add_employee", methods=["GET", "POST"])
def add_employee():


    if 'role' not in session or session['role'] != 'admin':
        return "Access Denied"



    if request.method == "POST":


        # Get form data

        name = request.form.get("name")

        email = request.form.get("email")

        password = request.form.get("password")

        role = request.form.get("role", "staff")
        position = request.form.get("position", "").strip()
        can_approve_requests = request.form.get("can_approve_requests") == "1"
        immediate_supervisor_id = request.form.get("immediate_supervisor_id") or None



        # Validation

        if not name or not email or not password:

            return "Please fill all fields", 400



        hashed_password = generate_password_hash(password)



        conn = get_db()

        c = conn.cursor()



        c.execute("""
            INSERT INTO employees
            (
                name,
                email,
                password,
                role,
                position,
                can_approve_requests,
                immediate_supervisor_id
            )

            VALUES
            (
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s
            )
        """,
        (
            name,
            email,
            hashed_password,
            role,
            position,
            can_approve_requests,
            immediate_supervisor_id
        ))



        conn.commit()

        conn.close()



        return redirect("/admin/employees")



    return render_template(
        "add_employee.html"
    )

@app.route('/admin/save_employee', methods=['POST'])
def save_employee():

    if session.get('role') != 'admin':
        return jsonify({
            "success": False,
            "message": "Access Denied"
        }), 403

    employee_id = request.form.get('employee_id', '').strip()
    name = request.form.get('name', '').strip()
    email = request.form.get('email', '').strip()
    phone = request.form.get('phone', '').strip()
    department = request.form.get('department', '').strip()
    position = request.form.get('position', '').strip()
    can_approve_requests = request.form.get('can_approve_requests') == '1'
    immediate_supervisor_id = request.form.get('immediate_supervisor_id') or None
    role = request.form.get('role', 'employee').strip().lower()
    password = request.form.get('password', '').strip()

    if not name or not email:
        return jsonify({
            "success": False,
            "message": "Name and email are required."
        }), 400

    if role not in ('admin', 'employee', 'staff'):
        role = 'employee'

    conn = get_db()
    c = conn.cursor(cursor_factory=RealDictCursor)

    try:
        # -------------------------------------------------
        # EDIT EXISTING EMPLOYEE
        # -------------------------------------------------
        if employee_id:

            try:
                employee_id = int(employee_id)
            except ValueError:
                return jsonify({
                    "success": False,
                    "message": "Invalid employee ID."
                }), 400

            c.execute("""
                SELECT id
                FROM employees
                WHERE id=%s
            """, (employee_id,))

            if not c.fetchone():
                return jsonify({
                    "success": False,
                    "message": "Employee not found."
                }), 404

            # Do not replace the password when edit form is blank.
            if password:
                hashed_password = generate_password_hash(password)

                c.execute("""
                    UPDATE employees
                    SET
                        name=%s,
                        email=%s,
                        phone=%s,
                        department=%s,
                        position=%s,
                        can_approve_requests=%s,
                        immediate_supervisor_id=%s,
                        role=%s,
                        password=%s
                    WHERE id=%s
                """, (
                    name,
                    email,
                    phone,
                    department,
                    position,
                    can_approve_requests,
                    immediate_supervisor_id,
                    role,
                    hashed_password,
                    employee_id
                ))
            else:
                c.execute("""
                    UPDATE employees
                    SET
                        name=%s,
                        email=%s,
                        phone=%s,
                        department=%s,
                        position=%s,
                        can_approve_requests=%s,
                        immediate_supervisor_id=%s,
                        role=%s
                    WHERE id=%s
                """, (
                    name,
                    email,
                    phone,
                    department,
                    position,
                    can_approve_requests,
                    immediate_supervisor_id,
                    role,
                    employee_id
                ))

            conn.commit()

            # If an admin edited their own name, keep the session current.
            if employee_id == session.get('user_id'):
                session['name'] = name
                session['role'] = role

            return jsonify({
                "success": True,
                "message": f"{name}'s profile was updated successfully."
            })

        # -------------------------------------------------
        # CREATE NEW EMPLOYEE
        # -------------------------------------------------
        if not password:
            return jsonify({
                "success": False,
                "message": "Password is required when creating an employee."
            }), 400

        c.execute("""
            SELECT id
            FROM employees
            WHERE LOWER(email)=LOWER(%s)
        """, (email,))

        if c.fetchone():
            return jsonify({
                "success": False,
                "message": "An employee with that email already exists."
            }), 409

        hashed_password = generate_password_hash(password)

        c.execute("""
            INSERT INTO employees
            (
                name,
                email,
                password,
                role,
                phone,
                department,
                position,
                can_approve_requests,
                immediate_supervisor_id,
                theme
            )
            VALUES
            (
                %s,%s,%s,%s,%s,%s,%s,%s,%s,%s
            )
            RETURNING id
        """, (
            name,
            email,
            hashed_password,
            role,
            phone,
            department,
            position,
            can_approve_requests,
            immediate_supervisor_id,
            'light'
        ))

        new_id = c.fetchone()['id']

        conn.commit()

        return jsonify({
            "success": True,
            "message": f"{name} was added successfully.",
            "employee_id": new_id
        })

    except Exception as e:
        conn.rollback()
        print("SAVE EMPLOYEE ERROR:", repr(e))

        return jsonify({
            "success": False,
            "message": "Unable to save employee."
        }), 500

    finally:
        conn.close()


@app.route('/admin/edit_employee/<int:id>', methods=['GET', 'POST'])
def edit_employee(id):

    if session.get('role') != 'admin':
        return "Access Denied", 403

    conn = get_db()
    c = conn.cursor(cursor_factory=RealDictCursor)

    try:

        if request.method == 'POST':

            name = request.form.get('name', '').strip()
            email = request.form.get('email', '').strip()
            phone = request.form.get('phone', '').strip()
            department = request.form.get('department', '').strip()
            position = request.form.get('position', '').strip()
            can_approve_requests = request.form.get('can_approve_requests') == '1'
            immediate_supervisor_id = request.form.get('immediate_supervisor_id') or None

            role = request.form.get(
                'role',
                'staff'
            ).strip().lower()

            if not name or not email:
                return "Name and email are required.", 400

            # Only two roles are allowed.
            if role not in ('admin', 'staff'):
                role = 'staff'

            c.execute("""
                UPDATE employees
                SET
                    name=%s,
                    email=%s,
                    phone=%s,
                    department=%s,
                    position=%s,
                    can_approve_requests=%s,
                    immediate_supervisor_id=%s,
                    role=%s
                WHERE id=%s
            """, (
                name,
                email,
                phone,
                department,
                position,
                can_approve_requests,
                immediate_supervisor_id,
                role,
                id
            ))

            conn.commit()

            return redirect('/admin/employees')


        c.execute("""
            SELECT
                id,
                name,
                email,
                role,
                phone,
                department,
                position,
                can_approve_requests,
                immediate_supervisor_id,
                signature_path,
                profile_pic,
                theme,
                profile_pic_public_id
            FROM employees
            WHERE id=%s
        """, (id,))

        employee = c.fetchone()

        if not employee:
            return "Employee not found.", 404

        return render_template(
            "edit_employee.html",
            employee=employee
        )

    finally:
        conn.close()


@app.route('/admin/reset_password/<int:id>', methods=['POST'])
def reset_employee_password(id):

    if session.get('role') != 'admin':
        return jsonify({
            "success": False,
            "message": "Access Denied"
        }), 403

    password = request.form.get('password', '').strip()

    if len(password) < 4:
        return jsonify({
            "success": False,
            "message": "Password must be at least 4 characters."
        }), 400

    conn = get_db()
    c = conn.cursor(cursor_factory=RealDictCursor)

    try:
        c.execute("""
            SELECT id, name
            FROM employees
            WHERE id=%s
        """, (id,))

        employee = c.fetchone()

        if not employee:
            return jsonify({
                "success": False,
                "message": "Employee not found."
            }), 404

        hashed_password = generate_password_hash(password)

        c.execute("""
            UPDATE employees
            SET password=%s
            WHERE id=%s
        """, (
            hashed_password,
            id
        ))

        conn.commit()

        return jsonify({
            "success": True,
            "message": f"Password for {employee['name']} has been reset."
        })

    except Exception as e:
        conn.rollback()
        print("RESET PASSWORD ERROR:", repr(e))

        return jsonify({
            "success": False,
            "message": "Could not reset password."
        }), 500

    finally:
        conn.close()


@app.route('/admin/delete_employee/<int:id>', methods=['POST'])
def delete_employee(id):

    if session.get('role') != 'admin':
        return jsonify({
            "success": False,
            "message": "Access Denied"
        }), 403

    # Prevent an administrator from deleting their own account.
    if id == session.get('user_id'):
        return jsonify({
            "success": False,
            "message": "You cannot delete your own account."
        }), 400

    conn = get_db()
    c = conn.cursor(cursor_factory=RealDictCursor)

    try:

        # -----------------------------------------
        # Check employee exists
        # -----------------------------------------

        c.execute("""
            SELECT id, name
            FROM employees
            WHERE id=%s
        """, (id,))

        employee = c.fetchone()

        if not employee:
            return jsonify({
                "success": False,
                "message": "Employee not found."
            }), 404


        # -----------------------------------------
        # Delete attendance records
        # -----------------------------------------

        c.execute("""
            DELETE FROM attendance
            WHERE employee_id=%s
        """, (id,))


        # -----------------------------------------
        # Delete task comments
        # -----------------------------------------

        c.execute("""
            DELETE FROM task_comments
            WHERE sender_id=%s
        """, (id,))


        # -----------------------------------------
        # Delete assigned tasks
        # -----------------------------------------

        c.execute("""
            DELETE FROM tasks
            WHERE assigned_to=%s
        """, (id,))


        # -----------------------------------------
        # Delete notifications
        # -----------------------------------------

        c.execute("""
            DELETE FROM notifications
            WHERE user_id=%s
        """, (id,))


        # -----------------------------------------
        # Delete announcements created by employee
        # -----------------------------------------

        c.execute("""
            DELETE FROM announcements
            WHERE created_by=%s
        """, (id,))


        c.execute("""
            DELETE FROM messages
            WHERE sender_id=%s
        """, (id,))

        # -----------------------------------------
        # Finally delete employee
        # -----------------------------------------

        c.execute("""
            DELETE FROM employees
            WHERE id=%s
        """, (id,))


        conn.commit()


        return jsonify({
            "success": True,
            "message": f"{employee['name']} was deleted successfully."
        })


    except Exception as e:

        conn.rollback()

        print(
            "DELETE EMPLOYEE ERROR:",
            repr(e)
        )

        return jsonify({
            "success": False,
            "message": "Could not delete employee."
        }), 500


    finally:

        conn.close()


# =====================================================
# ANNOUNCEMENT DIRECT CLOUDINARY UPLOAD
# =====================================================

ANNOUNCEMENT_MAX_FILE_SIZE = 10 * 1024 * 1024

ANNOUNCEMENT_ALLOWED_EXTENSIONS = {
    'pdf', 'doc', 'docx', 'xls', 'xlsx',
    'ppt', 'pptx', 'txt', 'csv',
    'jpg', 'jpeg', 'png', 'gif', 'webp'
}


def _announcement_admin_required():
    return (
        'user_id' in session
        and session.get('role') == 'admin'
    )


def _cloudinary_config_values():
    config = cloudinary.config()

    return (
        config.cloud_name,
        config.api_key,
        config.api_secret
    )


def _delete_cloudinary_asset(public_id, resource_type):
    """
    Best-effort deletion of an announcement asset.

    New direct-upload announcements store the public_id and resource_type,
    so cleanup does not need to guess from a delivery URL.
    """
    if not public_id:
        return

    try:
        uploader.destroy(
            public_id,
            resource_type=resource_type or 'image',
            type='upload',
            invalidate=True
        )
    except Exception as exc:
        print(
            "CLOUDINARY ANNOUNCEMENT CLEANUP ERROR:",
            repr(exc)
        )


def _verify_cloudinary_asset(
    public_id,
    resource_type,
    secure_url,
    original_name
):
    """
    Verify the uploaded Cloudinary asset server-side.

    The browser never sends the file through Flask, so the finalization
    endpoint verifies the actual Cloudinary asset metadata before the
    announcement is saved.
    """

    if not public_id or not secure_url:
        raise ValueError(
            "Cloudinary did not return a valid uploaded asset."
        )

    cloud_name, _, _ = _cloudinary_config_values()

    expected_prefix = (
        f"https://res.cloudinary.com/{cloud_name}/"
    )

    if not secure_url.startswith(expected_prefix):
        raise ValueError(
            "Invalid Cloudinary asset URL."
        )

    extension = (
        secure_filename(original_name or "")
        .rsplit('.', 1)[1].lower()
        if '.' in secure_filename(original_name or "")
        else ''
    )

    if extension not in ANNOUNCEMENT_ALLOWED_EXTENSIONS:
        raise ValueError(
            "Invalid attachment type."
        )

    asset = cloudinary_api.resource(
        public_id,
        resource_type=resource_type or 'image',
        type='upload'
    )

    actual_bytes = int(
        asset.get('bytes') or 0
    )

    if actual_bytes > ANNOUNCEMENT_MAX_FILE_SIZE:

        _delete_cloudinary_asset(
            public_id,
            resource_type
        )

        raise ValueError(
            "Attachment is too large. Maximum allowed size is 10 MB."
        )

    actual_format = (
        str(asset.get('format') or '').lower()
    )

    # Cloudinary may represent some document formats differently,
    # so the original extension remains the primary allow-list check.
    if not actual_format:
        raise ValueError(
            "Cloudinary could not verify the attachment format."
        )

    return {
        "bytes": actual_bytes,
        "format": actual_format,
        "public_id": public_id,
        "resource_type": resource_type or 'image',
        "secure_url": secure_url,
        "asset_id": asset.get('asset_id')
    }


@app.route(
    '/admin/announcements/upload-signature',
    methods=['POST']
)
def announcement_upload_signature():

    if not _announcement_admin_required():
        return jsonify({
            "status": "error",
            "message": "Access denied."
        }), 403

    data = request.get_json(
        silent=True
    ) or {}

    audience = str(
        data.get('audience') or 'everyone'
    ).strip().lower()

    if audience not in (
        'everyone',
        'staff',
        'admin'
    ):
        audience = 'everyone'

    cloud_name, api_key, api_secret = (
        _cloudinary_config_values()
    )

    if not cloud_name or not api_key or not api_secret:
        return jsonify({
            "status": "error",
            "message": "Cloudinary is not configured on the server."
        }), 500

    # Use a SIGNED upload preset so Cloudinary itself can enforce
    # the 10 MB limit and allowed formats.
    upload_preset = os.environ.get(
        "CLOUDINARY_ANNOUNCEMENT_UPLOAD_PRESET"
    )

    if not upload_preset:
        return jsonify({
            "status": "error",
            "message": (
                "CLOUDINARY_ANNOUNCEMENT_UPLOAD_PRESET "
                "is not configured."
            )
        }), 500

    timestamp = int(
        time.time()
    )

    folder = (
        f"salt_portal/announcements/{audience}"
    )

    # Cloudinary requires the same signed fields to be sent unchanged
    # with the browser upload.
    params_to_sign = {
        "folder": folder,
        "timestamp": timestamp,
        "upload_preset": upload_preset
    }

    signature = cloudinary.utils.api_sign_request(
        params_to_sign,
        api_secret
    )

    return jsonify({
        "status": "success",
        "cloud_name": cloud_name,
        "api_key": api_key,
        "timestamp": timestamp,
        "signature": signature,
        "folder": folder,
        "upload_preset": upload_preset,
        "resource_type": "auto",
        "max_file_size": ANNOUNCEMENT_MAX_FILE_SIZE
    })


@app.route(
    '/admin/announcements',
    methods=['GET', 'POST']
)
def admin_announcements():

    # -------------------------
    # ACCESS CONTROL
    # -------------------------
    if 'user_id' not in session:
        return redirect('/')

    if session.get('role') != 'admin':
        return "Access Denied", 403

    conn = get_db()
    c = conn.cursor(cursor_factory=RealDictCursor)

    # =====================================================
    # CREATE ANNOUNCEMENT
    # =====================================================
    if request.method == 'POST':

        # The new browser flow sends JSON only.
        # No multipart file is ever sent to Flask.
        payload = request.get_json(
            silent=True
        )

        if payload is None:
            conn.close()

            return jsonify({
                "status": "error",
                "message": (
                    "Please use the current announcement form. "
                    "File uploads now go directly to Cloudinary."
                )
            }), 400

        title = str(
            payload.get('title') or ''
        ).strip()

        message = str(
            payload.get('message') or ''
        ).strip()

        audience = str(
            payload.get('audience') or 'everyone'
        ).strip().lower()

        if not title:
            conn.close()

            return jsonify({
                "status": "error",
                "message": "Please enter an announcement title."
            }), 400

        if not message:
            conn.close()

            return jsonify({
                "status": "error",
                "message": "Please enter an announcement message."
            }), 400

        if audience not in (
            'everyone',
            'staff',
            'admin'
        ):
            audience = 'everyone'

        uploaded_asset = (
            payload.get('file')
            if isinstance(payload.get('file'), dict)
            else None
        )

        file_name = None
        file_path = None
        public_id = None
        resource_type = None
        asset_id = None

        # -------------------------
        # VERIFY CLOUDINARY ASSET
        # -------------------------
        if uploaded_asset:

            file_name = secure_filename(
                str(
                    uploaded_asset.get('name') or ''
                )
            )

            file_path = str(
                uploaded_asset.get('secure_url') or ''
            )

            public_id = str(
                uploaded_asset.get('public_id') or ''
            )

            resource_type = str(
                uploaded_asset.get('resource_type')
                or 'image'
            )

            try:

                verified = _verify_cloudinary_asset(
                    public_id,
                    resource_type,
                    file_path,
                    file_name
                )

                asset_id = verified.get(
                    "asset_id"
                )

            except Exception as exc:

                _delete_cloudinary_asset(
                    public_id,
                    resource_type
                )

                conn.close()

                print(
                    "ANNOUNCEMENT ASSET VERIFICATION ERROR:",
                    repr(exc)
                )

                return jsonify({
                    "status": "error",
                    "message": str(exc)
                }), 400

        # -------------------------
        # CREATE DATABASE RECORD
        # -------------------------
        now = datetime.now()

        try:

            c.execute("""
                INSERT INTO announcements
                (
                    title,
                    message,
                    created_by,
                    created_at,
                    updated_at,
                    audience,
                    file_name,
                    file_path,
                    cloudinary_public_id,
                    cloudinary_resource_type,
                    cloudinary_asset_id
                )
                VALUES
                (
                    %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s
                )
            """, (
                title,
                message,
                session['user_id'],
                now,
                now,
                audience,
                file_name,
                file_path,
                public_id,
                resource_type,
                asset_id
            ))

            conn.commit()

        except Exception as exc:

            conn.rollback()

            if public_id:
                _delete_cloudinary_asset(
                    public_id,
                    resource_type
                )

            conn.close()

            print(
                "CREATE ANNOUNCEMENT DATABASE ERROR:",
                repr(exc)
            )

            return jsonify({
                "status": "error",
                "message": (
                    "Announcement could not be saved."
                )
            }), 500

        conn.close()

        # Notifications happen in the background so the admin
        # receives the success response immediately.
        socketio.start_background_task(
            notify_announcement,
            audience,
            f"New Announcement: {title}",
            session['user_id']
        )

        return jsonify({
            "status": "success",
            "message": "Announcement published successfully."
        })

    # =====================================================
    # SEARCH
    # =====================================================
    search = request.args.get(
        'search',
        ''
    ).strip()

    # =====================================================
    # AUDIENCE FILTER
    # =====================================================
    audience_filter = request.args.get(
        'audience',
        'all'
    ).strip().lower()

    # Only accept valid filters
    if audience_filter not in (
        'all',
        'everyone',
        'staff',
        'admin'
    ):
        audience_filter = 'all'

    # =====================================================
    # BUILD ANNOUNCEMENT QUERY
    # =====================================================
    query = """
        SELECT
            announcements.id,
            announcements.title,
            announcements.message,
            announcements.created_by,
            announcements.created_at,
            announcements.updated_at,
            announcements.audience,
            announcements.file_name,
            announcements.file_path,

            COALESCE(
                employees.name,
                'Unknown'
            ) AS name

        FROM announcements

        LEFT JOIN employees
            ON announcements.created_by = employees.id

        WHERE 1=1
    """

    params = []

    # -------------------------
    # SEARCH FILTER
    # -------------------------
    if search:

        query += """
            AND (
                announcements.title ILIKE %s
                OR announcements.message ILIKE %s
                OR employees.name ILIKE %s
            )
        """

        search_value = f"%{search}%"

        params.extend([
            search_value,
            search_value,
            search_value
        ])

    # -------------------------
    # AUDIENCE FILTER
    # -------------------------
    if audience_filter != 'all':

        query += """
            AND announcements.audience=%s
        """

        params.append(
            audience_filter
        )

    # -------------------------
    # SORT
    # -------------------------
    query += """
        ORDER BY announcements.created_at DESC
    """

    c.execute(
        query,
        params
    )

    announcements = c.fetchall()

    # =====================================================
    # ANNOUNCEMENT STATISTICS
    # =====================================================

    # Total
    c.execute("""
        SELECT COUNT(*) AS count
        FROM announcements
    """)

    total_announcements = c.fetchone()['count']

    # Everyone
    c.execute("""
        SELECT COUNT(*) AS count
        FROM announcements
        WHERE audience='everyone'
           OR audience IS NULL
    """)

    everyone_count = c.fetchone()['count']

    # Staff
    c.execute("""
        SELECT COUNT(*) AS count
        FROM announcements
        WHERE audience='staff'
    """)

    staff_count = c.fetchone()['count']

    # Admin
    c.execute("""
        SELECT COUNT(*) AS count
        FROM announcements
        WHERE audience='admin'
    """)

    admin_count = c.fetchone()['count']

    conn.close()

    # =====================================================
    # RENDER PAGE
    # =====================================================
    return render_template(
        "admin_announcements.html",

        announcements=announcements,

        search=search,

        audience_filter=audience_filter,

        total_announcements=total_announcements,

        everyone_count=everyone_count,

        staff_count=staff_count,

        admin_count=admin_count
    )

# =====================================================
# DOWNLOAD ANNOUNCEMENT ATTACHMENT
# =====================================================

@app.route('/admin/announcement/file/<int:id>')
def download_announcement_file(id):

    if 'user_id' not in session:
        return redirect('/')

    conn = get_db()
    c = conn.cursor(
        cursor_factory=RealDictCursor
    )

    c.execute("""
        SELECT file_path, file_name
        FROM announcements
        WHERE id=%s
    """, (id,))

    announcement = c.fetchone()
    conn.close()

    if not announcement:
        return "Announcement not found", 404

    if not announcement['file_path']:
        return "No attachment", 404

    cloudinary_url = announcement['file_path']

    # View mode: preserve the original secure delivery URL.
    # Download mode: ask Cloudinary to force attachment delivery.
    if request.args.get("download") == "1":
        if "/upload/" in cloudinary_url:
            cloudinary_url = cloudinary_url.replace(
                "/upload/",
                "/upload/fl_attachment/"
            )

    return redirect(cloudinary_url)

# =====================================================
# UPDATE ANNOUNCEMENT
# =====================================================

@app.route(
    '/admin/update_announcement/<int:id>',
    methods=['POST']
)
def update_announcement(id):

    if 'user_id' not in session:
        return jsonify({
            "status": "error",
            "message": "Please log in."
        }), 401

    if session.get('role') != 'admin':
        return jsonify({
            "status": "error",
            "message": "Access denied."
        }), 403

    payload = request.get_json(
        silent=True
    )

    if payload is None:
        return jsonify({
            "status": "error",
            "message": (
                "Please use the current announcement editor."
            )
        }), 400

    title = str(
        payload.get('title') or ''
    ).strip()

    message = str(
        payload.get('message') or ''
    ).strip()

    audience = str(
        payload.get('audience') or 'everyone'
    ).strip().lower()

    remove_file = bool(
        payload.get('remove_file')
    )

    uploaded_asset = (
        payload.get('file')
        if isinstance(payload.get('file'), dict)
        else None
    )

    if not title:
        return jsonify({
            "status": "error",
            "message": "Announcement title is required."
        }), 400

    if not message:
        return jsonify({
            "status": "error",
            "message": "Announcement message is required."
        }), 400

    if audience not in (
        'everyone',
        'staff',
        'admin'
    ):
        audience = 'everyone'

    conn = get_db()
    c = conn.cursor(
        cursor_factory=RealDictCursor
    )

    try:

        c.execute("""
            SELECT
                id,
                title,
                message,
                audience,
                file_name,
                file_path,
                cloudinary_public_id,
                cloudinary_resource_type,
                cloudinary_asset_id
            FROM announcements
            WHERE id=%s
        """, (id,))

        announcement = c.fetchone()

        if not announcement:
            return jsonify({
                "status": "error",
                "message": "Announcement not found."
            }), 404

        old_public_id = (
            announcement['cloudinary_public_id']
        )

        old_resource_type = (
            announcement['cloudinary_resource_type']
            or 'image'
        )

        new_file_name = announcement['file_name']
        new_file_path = announcement['file_path']
        new_public_id = old_public_id
        new_resource_type = old_resource_type
        new_asset_id = announcement['cloudinary_asset_id']

        uploaded_public_id = None
        uploaded_resource_type = None

        # -------------------------
        # REPLACEMENT ATTACHMENT
        # -------------------------
        if uploaded_asset:

            new_file_name = secure_filename(
                str(
                    uploaded_asset.get('name') or ''
                )
            )

            new_file_path = str(
                uploaded_asset.get('secure_url') or ''
            )

            uploaded_public_id = str(
                uploaded_asset.get('public_id') or ''
            )

            uploaded_resource_type = str(
                uploaded_asset.get('resource_type')
                or 'image'
            )

            try:

                verified = _verify_cloudinary_asset(
                    uploaded_public_id,
                    uploaded_resource_type,
                    new_file_path,
                    new_file_name
                )

                new_public_id = verified['public_id']
                new_resource_type = verified['resource_type']
                new_asset_id = verified.get('asset_id')

            except Exception as exc:

                _delete_cloudinary_asset(
                    uploaded_public_id,
                    uploaded_resource_type
                )

                return jsonify({
                    "status": "error",
                    "message": str(exc)
                }), 400

        elif remove_file:

            new_file_name = None
            new_file_path = None
            new_public_id = None
            new_resource_type = None
            new_asset_id = None

        # -------------------------
        # UPDATE DATABASE
        # -------------------------
        c.execute("""
            UPDATE announcements
            SET
                title=%s,
                message=%s,
                audience=%s,
                updated_at=%s,
                file_name=%s,
                file_path=%s,
                cloudinary_public_id=%s,
                cloudinary_resource_type=%s,
                cloudinary_asset_id=%s
            WHERE id=%s
        """, (
            title,
            message,
            audience,
            datetime.now(),
            new_file_name,
            new_file_path,
            new_public_id,
            new_resource_type,
            new_asset_id,
            id
        ))

        conn.commit()

        # Delete the previous Cloudinary asset only after the
        # database has successfully switched to the new one.
        if (
            uploaded_asset
            and old_public_id
            and old_public_id != new_public_id
        ):
            _delete_cloudinary_asset(
                old_public_id,
                old_resource_type
            )

        if (
            remove_file
            and old_public_id
        ):
            _delete_cloudinary_asset(
                old_public_id,
                old_resource_type
            )

        socketio.start_background_task(
            notify_announcement,
            audience,
            f"Announcement Updated: {title}",
            session['user_id']
        )

        return jsonify({
            "status": "success",
            "message": "Announcement updated successfully."
        })

    except Exception as exc:

        conn.rollback()

        # If a newly uploaded asset was not committed, clean it up.
        if uploaded_public_id:
            _delete_cloudinary_asset(
                uploaded_public_id,
                uploaded_resource_type
            )

        print(
            "UPDATE ANNOUNCEMENT ERROR:",
            repr(exc)
        )

        return jsonify({
            "status": "error",
            "message": "Failed to update announcement."
        }), 500

    finally:

        conn.close()

# =====================================================
# DELETE ANNOUNCEMENT
# =====================================================

@app.route(
    '/admin/delete_announcement/<int:id>',
    methods=['POST']
)
def delete_announcement(id):

    if 'user_id' not in session:
        return jsonify({
            "status": "error",
            "message": "Unauthorized"
        }), 401

    if session.get('role') != 'admin':
        return jsonify({
            "status": "error",
            "message": "Access Denied"
        }), 403

    conn = get_db()
    c = conn.cursor(
        cursor_factory=RealDictCursor
    )

    try:

        c.execute("""
            SELECT
                id,
                title,
                cloudinary_public_id,
                cloudinary_resource_type
            FROM announcements
            WHERE id=%s
        """, (id,))

        announcement = c.fetchone()

        if not announcement:
            return jsonify({
                "status": "error",
                "message": "Announcement not found."
            }), 404

        c.execute("""
            DELETE FROM announcements
            WHERE id=%s
        """, (id,))

        conn.commit()

        if announcement['cloudinary_public_id']:
            _delete_cloudinary_asset(
                announcement['cloudinary_public_id'],
                announcement['cloudinary_resource_type']
            )

        return jsonify({
            "status": "success",
            "message": "Announcement deleted successfully."
        })

    except Exception as exc:

        conn.rollback()

        print(
            "DELETE ANNOUNCEMENT ERROR:",
            repr(exc)
        )

        return jsonify({
            "status": "error",
            "message": "Unable to delete announcement."
        }), 500

    finally:
        conn.close()

@app.route('/tasks')
def tasks():

    if 'user_id' not in session:
        return redirect('/')


    conn = get_db()

    c = conn.cursor(
        cursor_factory=RealDictCursor
    )


    today = datetime.now().date()


    # =====================================================
    # AUTO CARRY FORWARD OVERDUE TASKS
    # =====================================================

    c.execute("""

        SELECT

            id,
            deadline,
            status

        FROM tasks

        WHERE assigned_to = %s

        AND status <> 'Completed'

        AND COALESCE(
            employee_deleted,
            FALSE
        ) = FALSE

    """,
    (
        session['user_id'],
    ))


    pending_tasks = c.fetchall()


    for task in pending_tasks:

        deadline = task["deadline"]


        if deadline:

            if isinstance(
                deadline,
                datetime
            ):

                deadline = deadline.date()


            elif isinstance(
                deadline,
                date
            ):

                pass


            elif isinstance(
                deadline,
                str
            ):

                try:

                    deadline = datetime.strptime(
                        deadline[:10],
                        "%Y-%m-%d"
                    ).date()

                except (
                    ValueError,
                    TypeError
                ):

                    deadline = None


            else:

                deadline = None


            if (
                deadline
                and deadline < today
            ):

                new_deadline = (
                    deadline +
                    timedelta(days=7)
                )


                c.execute("""

                    UPDATE tasks

                    SET

                        deadline = %s,

                        carried_forward = TRUE

                    WHERE id = %s

                """,
                (
                    new_deadline,
                    task["id"]
                ))


    conn.commit()


    # =====================================================
    # LOAD EMPLOYEE TASKS
    # =====================================================

    c.execute("""

        SELECT

            id,
            title,
            description,
            note,
            reply,
            admin_reply,
            deadline,
            status,
            created_by,
            assigned_to,
            carried_forward,
            completed_at,
            created_at,
            admin_deleted,
            employee_deleted

        FROM tasks

        WHERE assigned_to = %s

        AND COALESCE(
            employee_deleted,
            FALSE
        ) = FALSE

        ORDER BY id DESC

    """,
    (
        session['user_id'],
    ))


    tasks_list = c.fetchall()


    # =====================================================
    # COMMENTS / DISCUSSIONS
    # =====================================================

    for task in tasks_list:

        c.execute("""

            SELECT *

            FROM task_comments

            WHERE task_id = %s

            AND visibility = 'public'

            ORDER BY created_at ASC

        """,
        (
            task["id"],
        ))


        comments = c.fetchall()


        task["comments"] = (
            build_comment_tree(
                comments
            )
        )


    conn.close()


    return render_template(

        "tasks.html",

        tasks=tasks_list,

        name=session["name"],

        role=session["role"]

    )

@app.route('/reply_task/<int:id>', methods=['POST'])
def reply_task(id):

    if 'user_id' not in session:
        return redirect('/')

    message = request.form["reply"]

    parent_comment_id = request.form.get(
        "parent_comment_id"
    )

    conn = get_db()

    c = conn.cursor()

    c.execute("""
        INSERT INTO task_comments
        (
            task_id,
            sender_id,
            sender_role,
            message,
            parent_comment_id,
            visibility,
            created_at
        )
        VALUES
        (
            %s,
            %s,
            %s,
            %s,
            %s,
            %s,
            %s
        )
    """,
    (
        id,
        session["user_id"],
        "employee",
        message,
        parent_comment_id,
        "public",
        datetime.now()
    ))

    # Find the task creator
    c.execute("""
        SELECT created_by, title
        FROM tasks
        WHERE id=%s
    """, (id,))

    task_owner = c.fetchone()

    conn.commit()
    conn.close()

    # Notify the task creator
    if task_owner and task_owner["created_by"] != session["user_id"]:

        create_notification(
            task_owner["created_by"],
            f"New reply on task: {task_owner['title']}"
        )

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":

        return jsonify({
            "status": "success"
        })

    return redirect("/tasks")

@app.route('/task/add_note/<int:id>', methods=['POST'])
def add_task_note(id):

    if 'user_id' not in session:
        return redirect('/')

    note = request.form["note"]

    conn = get_db()

    c = conn.cursor()

    c.execute("""

        INSERT INTO task_comments
        (
            task_id,
            sender_id,
            sender_role,
            message,
            visibility,
            comment_type,
            created_at
        )

        VALUES
        (
            %s,
            %s,
            %s,
            %s,
            %s,
            %s,
            %s
        )

    """,
    (
        id,
        session["user_id"],
        session["role"],
        note,
        "public",
        "note",
        datetime.now()
    ))

    conn.commit()

    conn.close()

    return redirect("/tasks")

@app.route('/task-history')
def task_history():

    if 'user_id' not in session:
        return redirect('/')

    filter_type = request.args.get("filter", "all")

    search = request.args.get("search", "")

    conn = get_db()

    c = conn.cursor(cursor_factory=RealDictCursor)

    query = """
        SELECT *

        FROM tasks

        WHERE assigned_to=%s

        AND status='Completed'
    """

    params = [session["user_id"]]

    # ==========================
    # SEARCH
    # ==========================

    if search:

        query += """

            AND
            (
                title ILIKE %s
                OR description ILIKE %s
            )

        """

        params.extend([
            f"%{search}%",
            f"%{search}%"
        ])

    # ==========================
    # FILTERS
    # ==========================

    if filter_type == "this_week":

        query += """
            AND completed_at >=
            date_trunc('week', CURRENT_DATE)
        """

    elif filter_type == "last_week":

        query += """
            AND completed_at >=
                date_trunc('week', CURRENT_DATE)
                - interval '1 week'

            AND completed_at <
                date_trunc('week', CURRENT_DATE)
        """

    elif filter_type == "this_month":

        query += """
            AND date_part(
                'month',
                completed_at
            )
            =
            date_part(
                'month',
                CURRENT_DATE
            )

            AND date_part(
                'year',
                completed_at
            )
            =
            date_part(
                'year',
                CURRENT_DATE
            )
        """

    elif filter_type == "last_month":

        query += """
            AND date_part(
                'month',
                completed_at
            )
            =
            date_part(
                'month',
                CURRENT_DATE - interval '1 month'
            )

            AND date_part(
                'year',
                completed_at
            )
            =
            date_part(
                'year',
                CURRENT_DATE - interval '1 month'
            )
        """

    query += """

        ORDER BY completed_at DESC

    """

    c.execute(query, params)

    tasks = c.fetchall()

    conn.close()

    return render_template(
        "task_history.html",
        tasks=tasks,
        current_filter=filter_type,
        search=search
    )

@app.route('/export-task-history')
def export_task_history():

    if 'user_id' not in session:
        return redirect('/')

    export_type = request.args.get('type', 'all')

    conn = get_db()
    c = conn.cursor(cursor_factory=RealDictCursor)

    query = """
        SELECT
            id,
            title,
            description,
            deadline,
            status,
            completed_at,
            carried_forward
        FROM tasks
        WHERE assigned_to=%s
        AND status='Completed'
    """

    params = [session['user_id']]

    if export_type == "my":

        query += """
            AND created_by = assigned_to
        """

    elif export_type == "assigned":

        query += """
            AND created_by <> assigned_to
        """

    query += """
        ORDER BY completed_at DESC
    """

    c.execute(query, params)

    rows = c.fetchall()

    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Task History"

    ws.append([
        "ID",
        "Title",
        "Description",
        "Deadline",
        "Status",
        "Completed At",
        "Carried Forward"
    ])

    for row in rows:

        ws.append([
            row["id"],
            row["title"],
            row["description"],
            row["deadline"],
            row["status"],
            row["completed_at"],
            "Yes" if row["carried_forward"] else "No"
        ])

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 18

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    if export_type == "my":
        filename = "my_tasks.xlsx"

    elif export_type == "assigned":
        filename = "assigned_tasks.xlsx"

    else:
        filename = "all_tasks.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/complete_task/<int:id>')
def complete_task(id):

    if 'user_id' not in session:
        return redirect('/')

    user_id = session['user_id']

    conn = get_db()

    c = conn.cursor(
        cursor_factory=RealDictCursor
    )

    try:

        # ==========================================
        # GET TASK
        # ==========================================

        c.execute("""
            SELECT
                id,
                title,
                created_by,
                assigned_to,
                status
            FROM tasks
            WHERE id = %s
        """, (id,))

        task = c.fetchone()


        if not task:

            conn.close()

            flash(
                "Task not found.",
                "error"
            )

            return redirect('/tasks')


        # ==========================================
        # CHECK OWNERSHIP
        # ==========================================
        #
        # A user can complete:
        #
        # 1. A task assigned to them
        # 2. A self-created task
        #
        # This also supports your current
        # self-task structure.
        # ==========================================

        assigned_to = task.get(
            "assigned_to"
        )

        created_by = task.get(
            "created_by"
        )


        if (
            assigned_to != user_id
            and created_by != user_id
        ):

            conn.close()

            flash(
                "You are not authorized to complete this task.",
                "error"
            )

            return redirect('/tasks')


        # ==========================================
        # ALREADY COMPLETED
        # ==========================================

        if task.get("status") == "Completed":

            conn.close()

            flash(
                "This task is already completed.",
                "info"
            )

            return redirect('/tasks')


        # ==========================================
        # COMPLETE TASK
        # ==========================================

        c.execute("""
            UPDATE tasks
            SET
                status = 'Completed',
                completed_at = %s
            WHERE id = %s
        """, (
            datetime.now(),
            id
        ))


        # ==========================================
        # COMMIT
        # ==========================================

        conn.commit()


        # ==========================================
        # NOTIFY CREATOR
        # ==========================================
        #
        # Only notify if another employee completed
        # the task.
        #
        # Self-created task:
        # created_by == user_id
        # → no notification needed.
        # ==========================================

        if (
            created_by
            and created_by != user_id
        ):

            try:

                create_notification(
                    created_by,
                    f"Task Completed: {task['title']}"
                )

            except Exception as notification_error:

                print(
                    "TASK COMPLETION NOTIFICATION ERROR:",
                    notification_error
                )


        conn.close()


        flash(
            "Task completed successfully.",
            "success"
        )

        return redirect('/tasks')


    except Exception as e:

        conn.rollback()

        conn.close()

        print(
            "COMPLETE TASK ERROR:",
            e
        )

        flash(
            "Unable to complete task.",
            "error"
        )

        return redirect('/tasks')

@app.route('/tasks/create', methods=['POST'])
def create_task():

    if 'user_id' not in session:
        return redirect('/')

    title = request.form.get(
        "title",
        ""
    ).strip()

    description = request.form.get(
        "description",
        ""
    ).strip()

    deadline = request.form.get(
        "deadline"
    ) or None

    conn = get_db()

    c = conn.cursor()

    c.execute("""
        INSERT INTO tasks
        (
            title,
            description,
            assigned_to,
            deadline,
            status,
            created_by,
            created_at,
            original_deadline,
            carried_forward
        )
        VALUES
        (
            %s,
            %s,
            %s,
            %s,
            %s,
            %s,
            %s,
            %s,
            %s
        )
    """, (
        title,
        description,
        session["user_id"],
        deadline,
        "Pending",
        session["user_id"],
        datetime.now(),
        deadline,
        False
    ))

    conn.commit()
    conn.close()

    flash(
        "Task created successfully.",
        "success"
    )

    return redirect("/tasks")

@app.route('/start_task/<int:id>')
def start_task(id):

    if 'user_id' not in session:
        return redirect('/')

    conn = get_db()

    c = conn.cursor()

    c.execute("""
        SELECT created_by, title
        FROM tasks
        WHERE id=%s
        AND assigned_to=%s
    """, (
        id,
        session['user_id']
    ))

    task_owner = c.fetchone()

    c.execute("""
        UPDATE tasks
        SET status='In Progress'
        WHERE id=%s
        AND assigned_to=%s
    """, (
        id,
        session['user_id']
    ))

    conn.commit()

    conn.close()

    # 🔔 Notify the task creator when an assigned task is started.
    if task_owner and task_owner["created_by"] != session["user_id"]:

        create_notification(
            task_owner["created_by"],
            f"Task Started: {task_owner['title']}"
        )

    return redirect('/tasks')


@app.route('/delete_task/<int:id>')
def delete_task(id):

    if 'user_id' not in session:
        return redirect('/')

    conn = get_db()

    c = conn.cursor()

    c.execute("""

        DELETE FROM tasks

        WHERE id=%s

    """,
    (
        id,
    ))

    conn.commit()

    conn.close()

    return redirect('/tasks')

@app.route('/edit_task/<int:id>', methods=['POST'])
def edit_task(id):

    if 'user_id' not in session:
        return jsonify({
            "status":"unauthorized"
        }),403

    data = request.get_json()

    conn = get_db()

    c = conn.cursor()

    c.execute("""

        UPDATE tasks

        SET

            title=%s,

            description=%s

        WHERE id=%s

    """,
    (
        data["title"],
        data["description"],
        id
    ))

    conn.commit()

    conn.close()

    return jsonify({
        "status":"success"
    })

# SALTY AI API (FIXED + OPTIMIZED)
@app.route('/api/salty', methods=['POST'])
def salty_ai():

    if 'user_id' not in session:
        return jsonify({"reply": "Unauthorized"}), 403

    data = request.get_json() or {}
    msg = data.get("message", "").lower().strip()

    conn = get_db()
    c = conn.cursor(cursor_factory=RealDictCursor)

    today = datetime.now().date()

    # -------------------------
    # CORE DATA
    # -------------------------

    c.execute("SELECT COUNT(*) AS total FROM employees")
    total = c.fetchone()["total"]

    c.execute("""
        SELECT COUNT(DISTINCT employee_id) AS present
        FROM attendance
        WHERE DATE(clock_in)=%s
    """, (today,))

    present = c.fetchone()["present"]

    absent = total - present

    c.execute("""
        SELECT e.name
        FROM employees e
        JOIN attendance a
            ON e.id = a.employee_id
        WHERE DATE(a.clock_in)=%s
        AND a.clock_out IS NULL
    """, (today,))

    working = [row["name"] for row in c.fetchall()]

    conn.close()

    # -------------------------
    # Navigation
    # -------------------------

    nav_commands = {
        "open dashboard": "/dashboard",
        "dashboard": "/dashboard",
        "open attendance": "/attendance",
        "attendance": "/attendance",
        "open tasks": "/tasks",
        "tasks": "/tasks",
        "open employees": "/admin/employees",
        "employees": "/admin/employees",
        "logout": "/logout"
    }

    for cmd, route in nav_commands.items():

        if cmd in msg:

            return jsonify({
                "reply": f"Opening {cmd.replace('open ', '')}...",
                "action": route,
                "type": "navigation"
            })

    # -------------------------
    # Questions
    # -------------------------

    if "who is working" in msg:

        return jsonify({
            "reply": ", ".join(working) if working else "No one is currently working.",
            "type": "info"
        })

    if msg in ["working", "currently working"]:

        return jsonify({
            "reply": ", ".join(working) if working else "No active workers.",
            "type": "info"
        })

    if "present" in msg:

        return jsonify({
            "reply": f"{present} employees are present today.",
            "type": "stats"
        })

    if "absent" in msg:

        return jsonify({
            "reply": f"{absent} employees are absent today.",
            "type": "stats"
        })

    if "attendance" in msg:

        return jsonify({
            "reply": f"{present} present, {absent} absent today.",
            "type": "stats"
        })

    if "help" in msg:

        return jsonify({
            "reply": "Try: open dashboard, attendance, who is working, present, absent, open tasks",
            "type": "help"
        })

    return jsonify({
        "reply": "I can help with dashboard navigation, attendance, employees and tasks.",
        "type": "fallback"
    })


@app.route('/api/live-dashboard')
def live_dashboard():

    if 'user_id' not in session:
        return jsonify({
            "error":"unauthorized"
        }),403

    conn = get_db()

    c = conn.cursor(cursor_factory=RealDictCursor)

    today = datetime.now().date()

    c.execute("""
        SELECT COUNT(*) AS total
        FROM employees
    """)

    total = c.fetchone()["total"]

    c.execute("""
        SELECT COUNT(DISTINCT employee_id) AS present
        FROM attendance
        WHERE DATE(clock_in)=%s
    """,
    (
        today,
    ))

    present = c.fetchone()["present"]

    absent = total - present

    c.execute("""
        SELECT e.name

        FROM employees e

        JOIN attendance a

            ON e.id=a.employee_id

        WHERE DATE(a.clock_in)=%s

        AND a.clock_out IS NULL

        ORDER BY e.name
    """,
    (
        today,
    ))

    working = [
        row["name"]
        for row in c.fetchall()
    ]

    conn.close()

    return jsonify({

        "total":total,

        "present":present,

        "absent":absent,

        "working":working

    })

@socketio.on("join")
def on_join(data):

    if 'user_id' not in session:
        return

    expected_room = f"user_{session['user_id']}"
    requested_room = data.get("room")

    # Only allow a client to join its own notification room.
    if requested_room != expected_room:
        return

    join_room(expected_room)

    print(
        f"User {session['user_id']} joined notification room "
        f"{expected_room}"
    )

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/')
# -------------------------
# Initialize app
# -------------------------
try:
    init_db()
    create_admin()
    register_request_workflow(app, socketio)

    print("✅ Database initialized successfully.")

except Exception as e:
    print("\n==============================")
    print(" Application Startup Error")
    print("==============================")
    print(e)
    raise SystemExit(1)


if __name__ == "__main__":
    socketio.run(
        app,
        host="0.0.0.0",
        port=5000,
        debug=True
    )